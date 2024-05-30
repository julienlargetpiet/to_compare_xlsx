from openpyxl import load_workbook

from openpyxl.styles import Font

import openpyxl as openpyxl

###### CONF VAR #######

font_for_change = Font(name='Calibri',
                 size=9,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF5733')

lined = 5 #minimum is 1, it is the number of new cells that will be colored (relative number from the new column)

exception_l = [] # it is the exception cells when-it-comes to lined variable

###################

from_behavior = "no"

from_l = [1] #is the relative position of the columnn to see from the old column

##################

#######################

from_val = []

from_pre = []

colored = str(input("Do you want to color the new values? (y/n)"))

file_ = str(input("What is the file? "))

tab_ = str(input("What is the sheet? "))

row_ = int(input("Number of old row "))

col_ = str(input("Name of old column "))

col_ = int(openpyxl.utils.cell.column_index_from_string(col_))

row2_ = int(input("Number of new row "))

row2b_ = row2_

col2_ = str(input("Name of new column "))

col2_ = int(openpyxl.utils.cell.column_index_from_string(col2_))

col2b_ = col2_

print("")

a = load_workbook(file_)

sheet2 = a[tab_]

l_ = [] #old values

l2_ = [] #new values to compare to old values

e_sup = []

while sheet2.cell(row_, col_).value != None:
    
    val = str(sheet2.cell(row_, col_).value)

    if val[0] != "0":

        val = "0" + val

    l_.append(val)
  
    row_ += 1

while sheet2.cell(row2_, col2_).value != None:
    
    val = str(sheet2.cell(row2_, col2_).value)

    if val[0] != "0":

        val = "0" + val

    l2_.append(val)
  
    row2_ += 1

for i in range(0, len(l2_)):

    if l2_[i] not in l_:

        e_sup.append(l2_[i])
        
        if colored == "y":

            t = 0

            while sheet2.cell(row2b_ + i, col2_ + t).value != None and t < lined and t not in exception_l:

                sheet2.cell(row2b_ + i, col2_ + t).font = font_for_change

                t += 1

        from_pre.append(row2b_ + i)

if from_behavior == "yes":

    for fr in range (0, len(from_l)):

        for fr2  in range (0, len(from_pre)):

            from_val.append(sheet2.cell(from_pre[fr2], col2b_ + from_l[fr]).value)

print(from_val)
        
a.save(filename=file_)

print("Old values:", l_, len(l_))

print("")

print("New values:", l2_, len(l2_))

print("")

print("You have", len(e_sup), "new values. See them there!")

print(e_sup)        




